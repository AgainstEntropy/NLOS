# -*- coding: utf-8 -*-
# @Date    : 2022/4/30 15:52
# @Author  : WangYihao
# @File    : train.py
import os
import time

import torch
import adabound
from openpyxl import load_workbook
from torch import optim, nn
from torch.utils.tensorboard import SummaryWriter
from torchvision import transforms

from data.dataset import create_loaders, MyDataset
from my_utils import models
from my_utils.utils import get_device, check_accuracy, save_model


def trainer(model, optimizer, scheduler, loss_fn, train_loader,
            check_fn, check_loaders, batch_step, save_dir, log_every=10, epochs=2, writer=None):
    """

    Args:
        batch_step (int):
        epochs (int):
        log_every (int): log info per log_every batches.
        writer :

    Returns:
        batch_step (int):
    """
    device = get_device(model)
    # batch_size = train_loader.batch_size
    check_loader_train = check_loaders['train']
    check_loader_val = check_loaders['val']
    iters = len(train_loader)
    max_val_acc = 0.75

    for epoch in range(1, epochs + 1):
        tic = time.time()
        for batch_idx, (X, Y) in enumerate(train_loader):
            batch_step += 1
            model.train()
            X = X.to(device, dtype=torch.float32)
            Y = Y.to(device, dtype=torch.int64)
            # print(X.device, model.device)
            scores = model(X)
            loss = loss_fn(scores, Y)
            if writer is not None:
                writer.add_scalar('Metric/loss', loss.item(), batch_step)
                writer.add_scalar('Hpara/lr', optimizer.param_groups[0]['lr'], batch_step)

            # back propagate
            optimizer.zero_grad()
            loss.backward()
            optimizer.step()
            scheduler.step(batch_step / iters)

            # check accuracy
            if batch_idx % log_every == 0:
                model.eval()
                train_acc = check_fn(model, check_loader_train, training=True)
                val_acc = check_fn(model, check_loader_val, training=True)
                if writer is not None:
                    writer.add_scalars('Metric/acc', {'train': train_acc, 'val': val_acc}, batch_step)
                print(f'Epoch: {epoch} [{batch_idx}/{iters}]\tLoss: {loss:.4f}\t'
                      f'Val acc: {100. * val_acc:.1f}%')
                if val_acc > max_val_acc:
                    max_val_acc = val_acc
                    save_model(model, optimizer, scheduler,
                               save_dir=save_dir, acc=100 * val_acc)

        print(f'====> Epoch: {epoch}\tTime: {time.time() - tic}s')

    return batch_step


def train_a_model(model_configs=None, train_configs=None, loader_settings=None):
    """
        Train a model from zero.
    """
    if train_configs is None:
        train_configs = {
            'log_dir': 'tb_logs',
            'dataset_dir': '/mnt/cfs/wangyh/blender/blank_wall/datasets',
            'epochs': 50,
            'device': 'cuda:0',
            'optim': 'Adam',
            'lr': 1e-4,
            'schedule': 'cosine_warm',
            'cos_T': 15,
            'cos_mul': 2,
            'cos_iters': 3,
            'momentum': 0.9,
            'weight_decay': 0.05,
        }

    if loader_settings is not None:
        loader_settings = {
            'fig_resize': (128, 64),
            'batch_size': 16,
            'dataset_ratio': (0.6, 0.2, 0.2)
        }
    # make dataset
    # mean, std = torch.tensor(0.1094), torch.tensor(0.3660)  # polygons_unfilled_64_3
    T = transforms.Compose([
        transforms.ToTensor(),
        transforms.Resize(loader_settings['fig_resize']),
        # transforms.Normalize(mean, std)
    ])

    loader_kwargs = {
        'batch_size': loader_settings['batch_size'],  # default:1
        'shuffle': True,  # default:False
        'num_workers': 2,  # default:0
        'pin_memory': True,  # default:False
        'drop_last': True,  # default:False
        'prefetch_factor': 4,  # default:2
        'persistent_workers': False  # default:False
    }
    full_dataset = MyDataset(ROOT=train_configs['dataset_dir'], reduce_mode='W', transform=T)
    train_loader, val_loader, test_loader = create_loaders(
        full_dataset,
        ratio=loader_settings['dataset_ratio'],
        kwargs=loader_kwargs
    )
    check_loaders = {
        'train': train_loader,
        'val': val_loader
    }

    # create model
    if model_configs is None:
        model_configs = {
            'kernel_size': 7,
            'depths': (4, 1),
            'dims': (64, 64)
        }
    model, optimizer, scheduler = [None] * 3
    # define model
    model = models.NLOS_Conv(in_chans=3, num_classes=5, kernel_size=model_configs['kernel_size'],
                             depths=model_configs['depths'], dims=model_configs['dims'])
    model = model.to(train_configs['device'])

    # define optimizer
    if train_configs['optim'] == 'Adam':
        optimizer = optim.Adam(params=[{'params': model.parameters(), 'initial_lr': train_configs['lr']}],
                               lr=train_configs['lr'],
                               weight_decay=train_configs['weight_decay'])
    elif train_configs['optim'] == 'AdamW':
        optimizer = optim.AdamW(model.parameters(),
                                lr=train_configs['lr'],
                                weight_decay=train_configs['weight_decay'])
    elif train_configs['optim'] == 'AdaBound':
        optimizer = adabound.AdaBound(model.parameters(),
                                      lr=train_configs['lr'],
                                      weight_decay=train_configs['weight_decay'],
                                      final_lr=0.1)
    elif train_configs['optim'] == 'SGD':
        optimizer = optim.SGD(model.parameters(),
                              lr=train_configs['lr'],
                              weight_decay=train_configs['weight_decay'],
                              momentum=train_configs['momentum'])

    # define lr scheduler
    if train_configs['schedule'] == 'cosine_warm':
        train_configs['epochs'] = int((train_configs['cos_mul'] ** train_configs['cos_iters'] - 1) / \
                                      (train_configs['cos_mul'] - 1) * train_configs['cos_T'])
        scheduler = optim.lr_scheduler.CosineAnnealingWarmRestarts(optimizer,
                                                                   T_0=train_configs['cos_T'], T_mult=2)
    elif train_configs['schedule'] == 'cosine_anneal':
        scheduler = optim.lr_scheduler.CosineAnnealingLR(optimizer, T_max=train_configs['cos_T'])
    loss_func = nn.CrossEntropyLoss()
    print(f"model is on {next(model.parameters()).device}")

    # tensorboard writer
    save_time = time.strftime("%Y_%m_%d_%H_%M_%S", time.localtime())
    log_dir = os.path.join('/home/wangyh/01-Projects/01-NLOS/tb_logs',
                           f"{train_configs['log_dir']}")
    log_dir = os.path.join(log_dir, save_time)
    writer = SummaryWriter(log_dir=log_dir)
    # with open(os.path.join(log_dir, 'para.txt'), mode='w') as f:
    #     f.write('## -- model configs -- ##\n')
    #     for k, v in model_configs.items():
    #         f.write(f'{k} :\t{v}\n')
    #
    #     f.write('\n## -- dataset configs -- ##\n')
    #     for k, v in loader_kwargs.items():
    #         f.write(f'{k} :\t{v}\n')
    #
    #     f.write('\n## -- train configs -- ##\n')
    #     for k, v in train_configs.items():
    #         f.write(f'{k} :\t{v}\n')

    # record some configs
    # xlsx_path = '/home/wangyh/01-Projects/03-my/records/train_paras.xlsx'
    # wb = load_workbook(xlsx_path)
    # ws = wb['Sheet1']
    # record_data = [model_configs['type'],
    #                model_configs['kernel_size'],
    #                sum(model_configs['depths']),
    #                loader_kwargs['batch_size'],
    #                train_configs['lr'],
    #                train_configs['weight_decay'],
    #                train_configs['optim'],
    #                train_configs['schedule'],
    #                train_configs['cos_T'],
    #                train_configs['epochs']]
    # ws.append(record_data)
    # wb.save(filename=xlsx_path)
    # row = len(list(ws.values))

    save_dir = os.path.join(log_dir, 'weights')
    os.mkdir(save_dir)
    # start training!
    trainer(model=model, optimizer=optimizer,
            scheduler=scheduler,
            loss_fn=loss_func,
            train_loader=train_loader,
            check_fn=check_accuracy,
            check_loaders=check_loaders,
            batch_step=0, epochs=train_configs['epochs'], log_every=1,
            save_dir=save_dir,
            writer=writer)

    writer.close()

    final_acc = check_accuracy(model, test_loader, True)
    save_model(model, optimizer, scheduler, save_dir=save_dir, acc=100 * final_acc)

    # ws.cell(column=len(record_data) + 1, row=row, value=final_acc)
    # wb.save(filename=xlsx_path)


if __name__ == '__main__':
    model_configs = {
        'kernel_size': 5,
        'depths': (4, 1),
        'dims': (64, 64)
    }
    train_configs = {
        'log_dir': 'test_runs',
        'dataset_dir': '/mnt/cfs/wangyh/blender/blank_wall/datasets',
        'epochs': 50,
        'device': 'cuda:0',
        'optim': 'Adam',
        'lr': 1e-4,
        'schedule': 'cosine_warm',
        'cos_T': 15,
        'cos_mul': 2,
        'cos_iters': 3,
        'momentum': 0.9,
        'weight_decay': 0.05,
    }
    loader_settings = {
        'fig_resize': (128, 64),
        'batch_size': 8,
        'dataset_ratio': (0.6, 0.2, 0.2)
    }

    train_a_model(model_configs=model_configs, train_configs=train_configs, loader_settings=loader_settings)
